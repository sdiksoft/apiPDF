from flask import Flask, render_template, request, send_from_directory, jsonify, url_for
import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
from pathlib import Path
import subprocess
import platform
from weasyprint import HTML
from jinja2 import Template
import shutil
import threading
from queue import Queue
import time
import copy

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['DOWNLOAD_FOLDER'] = 'downloads'  # Pasta para PDFs gerados
app.config['TEMP_FOLDER'] = 'temp'  # Pasta para arquivos XLSX processados
app.config['MODEL_INFO'] = {}  # Armazena informações dos modelos
app.config['CONVERSION_QUEUE'] = Queue()  # Fila para conversão de PDFs
app.config['CONVERSION_STATUS'] = {}  # Status das conversões

def load_xlsx_models():
    """Carrega todos os modelos XLSX da pasta uploads ao iniciar a aplicação"""
    try:
        if not os.path.exists(app.config['UPLOAD_FOLDER']):
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            print("Pasta de uploads criada")
            return
        
        print("Carregando modelos XLSX...")
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            if filename.endswith('.xlsx'):
                try:
                    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                    wb = load_workbook(filepath)
                    sheet = wb.active
                    
                    # Inicializa as informações do modelo
                    model_info = {
                        'variables': [],
                        'tables': []
                    }
                    
                    # Procura por células com marcadores especiais
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                # Procura por variáveis (formato: ${nome:tipo})
                                var_matches = re.finditer(r'\${([^:]+):([^}]+)}', cell.value)
                                for match in var_matches:
                                    name, type_info = match.groups()
                                    model_info['variables'].append({
                                        'name': name,
                                        'type': type_info,
                                        'cell': cell.coordinate
                                    })
                                
                                # Procura por tabelas (formato: #{tabela.campo:tipo})
                                table_matches = re.finditer(r'#{([^.]+)\.([^:]+):([^}]+)}', cell.value)
                                for match in table_matches:
                                    table_name, field, type_info = match.groups()
                                    model_info['tables'].append({
                                        'name': table_name,
                                        'field': field,
                                        'type': type_info,
                                        'start_cell': cell.coordinate
                                    })
                                
                                # Procura por cálculos (formato: %{operacao})
                                calc_matches = re.finditer(r'%{([^}]+)}', cell.value)
                                for match in calc_matches:
                                    calc_expression = match.group(1)
                                    # Divide a expressão em tabela.campo:operação
                                    parts = calc_expression.split(':')
                                    if len(parts) != 2:
                                        continue
                                    
                                    field_parts = parts[0].split('.')
                                    if len(field_parts) != 2:
                                        continue
                                    
                                    table_name = field_parts[0]
                                    field_name = field_parts[1]
                                    operation = parts[1]
                                    
                                    model_info['calculations'] = model_info.get('calculations', [])
                                    model_info['calculations'].append({
                                        'table_name': table_name,
                                        'field_name': field_name,
                                        'operation': operation,
                                        'cell': cell.coordinate
                                    })
                    
                    # Armazena as informações do modelo
                    app.config['MODEL_INFO'][filename] = model_info
                    print(f"Modelo carregado: {filename}")
                    wb.close()
                
                except Exception as e:
                    print(f"Erro ao carregar modelo {filename}: {str(e)}")
        
        print("Carregamento de modelos concluído")
    
    except Exception as e:
        print(f"Erro ao carregar modelos: {str(e)}")

# Garante que as pastas necessárias existem
for folder in [app.config['UPLOAD_FOLDER'], app.config['DOWNLOAD_FOLDER'], app.config['TEMP_FOLDER']]:
    os.makedirs(folder, exist_ok=True)

# Carrega os modelos XLSX existentes
load_xlsx_models()

# Inicia o worker de conversão em background
def pdf_conversion_worker():
    while True:
        try:
            # Obtém o próximo item da fila
            excel_path, pdf_path = app.config['CONVERSION_QUEUE'].get()
            
            # Atualiza o status
            conversion_id = os.path.basename(pdf_path)
            app.config['CONVERSION_STATUS'][conversion_id] = {
                'status': 'processing',
                'message': 'Convertendo para PDF...'
            }
            
            try:
                # Verifica se o arquivo Excel existe e está pronto
                max_attempts = 5
                attempt = 0
                while attempt < max_attempts:
                    if os.path.exists(excel_path) and os.path.getsize(excel_path) > 0:
                        break
                    time.sleep(1)
                    attempt += 1
                
                if attempt >= max_attempts:
                    raise Exception("Arquivo Excel não está pronto para conversão")
                
                # Tenta converter usando LibreOffice
                if platform.system() == 'Linux':
                    # Verifica os possíveis caminhos do LibreOffice no Ubuntu
                    soffice_paths = [
                        '/usr/bin/soffice',
                        '/usr/bin/libreoffice',
                        '/usr/lib/libreoffice/program/soffice',
                    ]
                    
                    soffice = None
                    for path in soffice_paths:
                        if os.path.exists(path):
                            soffice = path
                            break
                    
                    if soffice:
                        # Mata qualquer processo do LibreOffice que possa estar rodando
                        try:
                            subprocess.run(['pkill', 'soffice'], stderr=subprocess.DEVNULL)
                            time.sleep(1)  # Espera um segundo para garantir que o processo foi finalizado
                        except:
                            pass
                        
                        # Garante que o diretório de destino existe
                        os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
                        
                        # Converte para PDF com tempo de espera maior
                        process = subprocess.run([
                            soffice,
                            '--headless',
                            '--convert-to', 'pdf:writer_pdf_Export',  # Usa o exportador PDF específico
                            '--outdir', os.path.dirname(pdf_path),
                            excel_path
                        ], check=True, timeout=60)  # Aumenta o timeout para 60 segundos
                        
                        # Espera um momento para o arquivo ser gerado
                        time.sleep(2)
                        
                        # Renomeia o arquivo
                        temp_pdf = os.path.join(os.path.dirname(pdf_path), 
                                              os.path.splitext(os.path.basename(excel_path))[0] + '.pdf')
                        
                        # Verifica se o PDF foi gerado e tem conteúdo
                        if os.path.exists(temp_pdf) and os.path.getsize(temp_pdf) > 0:
                            shutil.move(temp_pdf, pdf_path)
                            app.config['CONVERSION_STATUS'][conversion_id] = {
                                'status': 'completed',
                                'message': 'Conversão concluída com sucesso',
                                'pdf_url': f'/download/{os.path.basename(pdf_path)}'
                            }
                        else:
                            raise Exception("PDF não foi gerado corretamente")
                    else:
                        raise Exception("LibreOffice não encontrado. Instale com: sudo apt-get install libreoffice")
                else:
                    raise Exception("Sistema operacional não suportado")
                    
            except subprocess.TimeoutExpired:
                app.config['CONVERSION_STATUS'][conversion_id] = {
                    'status': 'error',
                    'message': 'Tempo limite excedido na conversão do PDF'
                }
            except Exception as e:
                app.config['CONVERSION_STATUS'][conversion_id] = {
                    'status': 'error',
                    'message': f'Erro na conversão: {str(e)}'
                }
            
            # Remove arquivos temporários após 1 hora
            threading.Timer(3600, cleanup_temp_files, args=[excel_path, pdf_path]).start()
            
        except Exception as e:
            print(f"Erro no worker de conversão: {str(e)}")
        finally:
            app.config['CONVERSION_QUEUE'].task_done()

def cleanup_temp_files(excel_path, pdf_path):
    """Remove arquivos temporários após um período"""
    try:
        if os.path.exists(excel_path):
            os.remove(excel_path)
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
    except Exception as e:
        print(f"Erro ao limpar arquivos temporários: {str(e)}")

# Inicia o thread do worker
conversion_thread = threading.Thread(target=pdf_conversion_worker, daemon=True)
conversion_thread.start()

@app.route('/conversion-status/<conversion_id>')
def conversion_status(conversion_id):
    """Retorna o status atual da conversão"""
    status = app.config['CONVERSION_STATUS'].get(conversion_id, {
        'status': 'not_found',
        'message': 'Conversão não encontrada'
    })
    return jsonify(status)

@app.route('/')
def index():
    # Lista todos os arquivos XLSX na pasta de uploads
    files = []
    if os.path.exists(app.config['UPLOAD_FOLDER']):
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            if filename.endswith('.xlsx'):
                model_info = app.config['MODEL_INFO'].get(filename, {})
                
                # Cria um exemplo de payload baseado nas informações do modelo
                example_payload = {}
                
                # Adiciona exemplos para variáveis simples
                for var in model_info.get('variables', []):
                    if var['type'] == 'text':
                        example_payload[var['name']] = f"Exemplo {var['name']}"
                    elif var['type'] == 'int':
                        example_payload[var['name']] = 123
                    elif var['type'] == 'double':
                        example_payload[var['name']] = 123.45
                    elif var['type'] == 'date':
                        example_payload[var['name']] = "11-03-2024"
                
                # Organiza campos por tabela
                tables = {}
                for table in model_info.get('tables', []):
                    table_name = table['name']
                    if table_name not in tables:
                        tables[table_name] = []
                    tables[table_name].append(table)
                
                # Adiciona exemplos para tabelas
                for table_name, fields in tables.items():
                    example_row = {}
                    for field in fields:
                        if field['type'] == 'text':
                            example_row[field['field']] = f"Exemplo {field['field']}"
                        elif field['type'] == 'int':
                            example_row[field['field']] = 123
                        elif field['type'] == 'double':
                            example_row[field['field']] = 123.45
                        elif field['type'] == 'date':
                            example_row[field['field']] = "11-03-2024"
                    
                    # Adiciona dois exemplos de linha para cada tabela
                    example_payload[table_name] = [
                        example_row,
                        {k: v for k, v in example_row.items()}  # Uma cópia do primeiro exemplo
                    ]
                
                file_info = {
                    'name': filename,
                    'endpoint': url_for('generate_from_model', model_name=filename[:-5], _external=True),
                    'schema': model_info,  # Mantém o schema original
                    'example_payload': example_payload  # Adiciona o exemplo de payload
                }
                files.append(file_info)
    return render_template('index.html', files=files)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return 'Nenhum arquivo enviado', 400
    
    file = request.files['file']
    if file.filename == '':
        return 'Nenhum arquivo selecionado', 400
    
    if not file.filename.endswith('.xlsx'):
        return 'Apenas arquivos XLSX são permitidos', 400

    try:
        # Garante que a pasta uploads existe
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        
        # Salva o arquivo
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(filepath)
        
        # Analisa o arquivo XLSX para extrair informações
        wb = load_workbook(filepath)
        sheet = wb.active
        
        # Inicializa as informações do modelo
        model_info = {
            'variables': [],
            'tables': []
        }
        
        # Procura por células com marcadores especiais
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Procura por variáveis (formato: ${nome:tipo})
                    var_matches = re.finditer(r'\${([^:]+):([^}]+)}', cell.value)
                    for match in var_matches:
                        name, type_info = match.groups()
                        model_info['variables'].append({
                            'name': name,
                            'type': type_info,
                            'cell': cell.coordinate
                        })
                    
                    # Procura por tabelas (formato: #{tabela.campo:tipo})
                    table_matches = re.finditer(r'#{([^.]+)\.([^:]+):([^}]+)}', cell.value)
                    for match in table_matches:
                        table_name, field, type_info = match.groups()
                        model_info['tables'].append({
                            'name': table_name,
                            'field': field,
                            'type': type_info,
                            'start_cell': cell.coordinate
                        })
                    
                    # Procura por cálculos (formato: %{tabela.campo:operação})
                    calc_matches = re.finditer(r'%{([^}]+)}', cell.value)
                    for match in calc_matches:
                        calc_expression = match.group(1)
                        # Divide a expressão em tabela.campo:operação
                        parts = calc_expression.split(':')
                        if len(parts) != 2:
                            continue
                        
                        field_parts = parts[0].split('.')
                        if len(field_parts) != 2:
                            continue
                        
                        table_name = field_parts[0]
                        field_name = field_parts[1]
                        operation = parts[1]
                        
                        model_info['calculations'] = model_info.get('calculations', [])
                        model_info['calculations'].append({
                            'table_name': table_name,
                            'field_name': field_name,
                            'operation': operation,
                            'cell': cell.coordinate
                        })
        
        # Armazena as informações do modelo
        app.config['MODEL_INFO'][file.filename] = model_info
        
        # Fecha o workbook
        wb.close()
        
        return 'Arquivo enviado com sucesso', 200
    
    except Exception as e:
        return f'Erro ao processar arquivo: {str(e)}', 500

@app.route('/download/<path:filename>')
def download_file(filename):
    # Verifica em todas as pastas possíveis
    for folder in [app.config['DOWNLOAD_FOLDER'], app.config['TEMP_FOLDER'], app.config['UPLOAD_FOLDER']]:
        if os.path.exists(os.path.join(folder, filename)):
            return send_from_directory(folder, filename)
    return "Arquivo não encontrado", 404

@app.route('/delete/<filename>', methods=['POST'])
def delete_model(filename):
    try:
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(file_path):
            os.remove(file_path)
            if filename in app.config['MODEL_INFO']:
                del app.config['MODEL_INFO'][filename]
            return jsonify({'message': 'Modelo excluído com sucesso'}), 200
        return jsonify({'error': 'Arquivo não encontrado'}), 404
    except Exception as e:
        return jsonify({'error': f'Erro ao excluir arquivo: {str(e)}'}), 500

@app.route('/api/generate/<model_name>', methods=['POST'])
def generate_from_model(model_name):
    filename = f'{model_name}.xlsx'
    if filename not in app.config['MODEL_INFO']:
        error_pdf = generate_error_pdf("Modelo não encontrado")
        if error_pdf:
            return jsonify({
                'error': 'Modelo não encontrado',
                'error_pdf': f'/download/{error_pdf}'
            }), 404
        return jsonify({'error': 'Modelo não encontrado'}), 404
    
    try:
        # Carrega o arquivo modelo
        template_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        wb = load_workbook(template_path)
        sheet = wb.active
        
        data = request.get_json()
        model_info = app.config['MODEL_INFO'][filename]
        
        # Substitui variáveis simples
        for var in model_info['variables']:
            if var['name'] in data:
                cell = sheet[var['cell']]
                value = data[var['name']]
                
                # Converte o valor para o tipo apropriado
                if var['type'] == 'date':
                    try:
                        value = datetime.strptime(value, '%d-%m-%Y')
                    except ValueError:
                        try:
                            value = datetime.strptime(value, '%Y-%m-%d')
                        except ValueError:
                            return jsonify({'error': f'Formato de data inválido para {var["name"]}. Use DD-MM-YYYY'}), 400
                elif var['type'] == 'int':
                    value = int(value)
                elif var['type'] == 'double':
                    value = float(value)
                
                cell.value = value
        
        # Organiza os campos por tabela e encontra células de cálculo
        tables = {}
        table_start_rows = {}
        table_columns = {}
        calculation_cells = {}  # Armazena as células que contêm cálculos
        
        # Primeiro, vamos encontrar todas as células de cálculo e suas informações
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # Suporta operações compostas como %[somar([produtos.quantidade.somar], [produtos.valor.media])]
                    cell_value = str(cell.value)
                    if '%[' in cell_value and ']' in cell_value:
                        # Extrai a expressão completa
                        full_expr = re.search(r'%\[(.*?)\]', cell_value)
                        if not full_expr:
                            continue
                        
                        calc_expression = full_expr.group(1)
                        
                        # Verifica se é uma operação composta
                        if '([' in calc_expression and '])' in calc_expression:
                            # Extrai a operação e as expressões
                            match = re.match(r'(\w+)\(\[(.*?)\],\s*\[(.*?)\]\)', calc_expression)
                            if match:
                                operation = match.group(1)
                                expr1 = match.group(2)
                                expr2 = match.group(3)
                                
                                col = ''.join(filter(str.isalpha, cell.coordinate))
                                row_num = int(''.join(filter(str.isdigit, cell.coordinate)))
                                calculation_cells[cell.coordinate] = {
                                    'type': 'compound',
                                    'operation': operation,
                                    'expr1': expr1,
                                    'expr2': expr2,
                                    'column': col,
                                    'row': row_num,
                                    'original_text': cell.value
                                }
                                continue
                        
                        # Se não for composta, processa como expressão simples
                        parts = calc_expression.split('.')
                        if len(parts) >= 3:
                            table_name = parts[0]
                            field_name = parts[1]
                            operation = parts[2]
                            
                            col = ''.join(filter(str.isalpha, cell.coordinate))
                            row_num = int(''.join(filter(str.isdigit, cell.coordinate)))
                            calculation_cells[cell.coordinate] = {
                                'type': 'simple',
                                'operation': operation,
                                'table_name': table_name,
                                'field_name': field_name,
                                'column': col,
                                'row': row_num,
                                'original_text': cell.value
                            }
        
        # Organiza as informações das tabelas
        for table_info in model_info.get('tables', []):
            table_name = table_info['name']
            if table_name not in tables:
                tables[table_name] = []
                start_row = int(''.join(filter(str.isdigit, table_info['start_cell'])))
                table_start_rows[table_name] = start_row
                table_columns[table_name] = {}
            
            col = ''.join(filter(str.isalpha, table_info['start_cell']))
            table_columns[table_name][table_info['field']] = col
            
            tables[table_name].append(table_info)
        
        # Processa cada tabela
        table_positions = {}  # Armazena a última linha usada para cada tabela
        
        for table_name, table_fields in tables.items():
            if table_name in data:
                table_data = data[table_name]
                if not isinstance(table_data, list):
                    return jsonify({'error': f'Dados da tabela {table_name} devem ser uma lista'}), 400
                
                start_row = table_start_rows[table_name]
                rows_to_insert = len(table_data)
                
                if rows_to_insert > 0:
                    # Encontra a última linha que contém cálculos para esta tabela
                    last_calc_row = start_row
                    for coord, calc_info in calculation_cells.items():
                        if calc_info['table_name'] == table_name:
                            row_num = int(''.join(filter(str.isdigit, coord)))
                            last_calc_row = max(last_calc_row, row_num)
                    
                    # Primeiro, salva os cálculos se existirem
                    calcs_to_move = {}
                    if last_calc_row > start_row:
                        for coord, calc_info in calculation_cells.items():
                            if calc_info['table_name'] == table_name:
                                row_num = int(''.join(filter(str.isdigit, coord)))
                                col = ''.join(filter(str.isalpha, coord))
                                if row_num == last_calc_row:
                                    calcs_to_move[col] = sheet[coord].value
                    
                    # Salva a formatação da linha modelo
                    template_row = start_row
                    template_formats = {}
                    max_col = 1
                    for cell in sheet[template_row]:
                        col_letter = cell.column_letter
                        max_col = max(max_col, cell.column)
                        template_formats[col_letter] = {
                            'font': copy.copy(cell.font),
                            'alignment': copy.copy(cell.alignment),
                            'border': copy.copy(cell.border),
                            'fill': copy.copy(cell.fill),
                            'number_format': cell.number_format,
                            'protection': copy.copy(cell.protection)
                        }
                    
                    # Insere as linhas necessárias
                    if rows_to_insert > 1:  # Só insere se precisar de mais de uma linha
                        sheet.insert_rows(start_row + 1, rows_to_insert - 1)
                    
                    # Para cada item na lista de dados
                    for idx, item in enumerate(table_data):
                        current_row = start_row + idx
                        
                        # Aplica a formatação na linha atual (exceto primeira linha que já está formatada)
                        if idx > 0:
                            for col in range(1, max_col + 1):
                                col_letter = get_column_letter(col)
                                if col_letter in template_formats:
                                    new_cell = sheet.cell(row=current_row, column=col)
                                    format_info = template_formats[col_letter]
                                    new_cell.font = copy.copy(format_info['font'])
                                    new_cell.alignment = copy.copy(format_info['alignment'])
                                    new_cell.border = copy.copy(format_info['border'])
                                    new_cell.fill = copy.copy(format_info['fill'])
                                    new_cell.number_format = format_info['number_format']
                                    new_cell.protection = copy.copy(format_info['protection'])
                        
                        # Para cada campo da tabela
                        for field in table_fields:
                            col = ''.join(filter(str.isalpha, field['start_cell']))
                            cell = f'{col}{current_row}'
                            
                            # Obtém e formata o valor
                            value = item.get(field['field'])
                            if value is not None:
                                if field['type'] == 'date':
                                    try:
                                        value = datetime.strptime(value, '%d-%m-%Y')
                                    except ValueError:
                                        try:
                                            value = datetime.strptime(value, '%Y-%m-%d')
                                        except ValueError:
                                            return jsonify({'error': f'Formato de data inválido para {field["field"]} em {table_name}'}), 400
                                elif field['type'] == 'int':
                                    value = int(value)
                                elif field['type'] == 'double':
                                    value = float(value)
                                    value = '{:.2f}'.format(value).replace('.', ',')
                                
                                # Atribui o valor mantendo a formatação
                                sheet[cell] = value
                                if field['type'] == 'double':
                                    sheet[cell].number_format = '#.##0,00'
                                elif field['type'] == 'int':
                                    sheet[cell].number_format = '#.##0'
                                elif field['type'] == 'date':
                                    sheet[cell].number_format = 'dd/mm/yyyy'
                    
                    # Move os cálculos para depois dos dados se existirem
                    if calcs_to_move:
                        new_calc_row = start_row + rows_to_insert
                        for col, formula in calcs_to_move.items():
                            new_coord = f"{col}{new_calc_row}"
                            old_coord = f"{col}{last_calc_row}"
                            if old_coord in calculation_cells:
                                calc_info = calculation_cells.pop(old_coord)
                                calc_info['row'] = new_calc_row
                                calculation_cells[new_coord] = calc_info
                            sheet[new_coord] = formula
                    # Atualiza table_positions para a próxima tabela
                    table_positions[table_name] = start_row + rows_to_insert - 1
        
        # Processa os cálculos após inserir todos os dados
        calculation_results = {}  # Armazena resultados intermediários
        
        # Primeiro, processa todos os cálculos simples
        for cell_coord, calc_info in calculation_cells.items():
            if calc_info['type'] == 'simple':
                result = None
                target_table = calc_info['table_name']
                target_column = None
                field_type = None
                
                # Encontra a coluna correta para o campo
                for table_info in tables[target_table]:
                    if table_info['field'] == calc_info['field_name']:
                        target_column = ''.join(filter(str.isalpha, table_info['start_cell']))
                        field_type = table_info['type']
                        break
                
                if target_column:
                    values = []
                    start_row = table_start_rows[target_table]
                    end_row = table_positions.get(target_table, start_row)
                    
                    for row in range(start_row, end_row + 1):
                        value_cell = sheet[f"{target_column}{row}"]
                        if value_cell.value is not None:
                            try:
                                value_str = str(value_cell.value)
                                if isinstance(value_cell.value, str):
                                    value = float(value_str.replace(',', '.'))
                                else:
                                    value = float(value_str)
                                values.append(value)
                            except (ValueError, TypeError):
                                pass
                    
                    if values:
                        if calc_info['operation'] == 'somar':
                            result = sum(values)
                        elif calc_info['operation'] == 'media':
                            result = sum(values) / len(values)
                        
                        # Armazena o resultado para uso em cálculos compostos
                        calculation_results[cell_coord] = result
                        
                        # Formata e exibe o resultado
                        if field_type == 'int':
                            result = int(result)
                        elif field_type == 'double':
                            result = '{:.2f}'.format(result).replace('.', ',')
                        
                        sheet[cell_coord].value = result
                        if field_type == 'double':
                            sheet[cell_coord].number_format = '#.##0,00'
        
        # Agora processa os cálculos compostos
        for cell_coord, calc_info in calculation_cells.items():
            if calc_info['type'] == 'compound':
                try:
                    # Encontra os resultados das expressões
                    def find_result(expr):
                        # Procura a célula que contém a expressão original
                        for coord, info in calculation_cells.items():
                            if info['type'] == 'simple':
                                expr_str = f"{info['table_name']}.{info['field_name']}.{info['operation']}"
                                if expr_str == expr:
                                    return calculation_results.get(coord)
                        return None
                    
                    result1 = find_result(calc_info['expr1'])
                    result2 = find_result(calc_info['expr2'])
                    
                    if result1 is not None and result2 is not None:
                        if calc_info['operation'] == 'somar':
                            result = result1 + result2
                        elif calc_info['operation'] == 'subtrair':
                            result = result1 - result2
                        elif calc_info['operation'] == 'multiplicar':
                            result = result1 * result2
                        elif calc_info['operation'] == 'dividir':
                            result = result1 / result2 if result2 != 0 else 0
                        
                        # Formata o resultado como número decimal
                        result_str = '{:.2f}'.format(result).replace('.', ',')
                        sheet[cell_coord].value = result_str
                        sheet[cell_coord].number_format = '#.##0,00'
                    else:
                        print(f"Não foi possível encontrar os resultados para: {calc_info['expr1']} ou {calc_info['expr2']}")
                except Exception as e:
                    print(f"Erro no cálculo composto: {str(e)}")
                    sheet[cell_coord].value = "ERRO"
        
        # Gera nomes únicos para os arquivos
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f'generated_{model_name}_{timestamp}.xlsx'
        pdf_filename = f'generated_{model_name}_{timestamp}.pdf'
        
        # Garante que as pastas existem
        os.makedirs(app.config['TEMP_FOLDER'], exist_ok=True)
        os.makedirs(app.config['DOWNLOAD_FOLDER'], exist_ok=True)
        
        # Salva o arquivo Excel temporário e força a sincronização
        excel_path = os.path.join(app.config['TEMP_FOLDER'], excel_filename)
        wb.save(excel_path)
        wb.close()
        
        # Força a sincronização do sistema de arquivos
        if platform.system() == 'Linux':
            try:
                subprocess.run(['sync'], check=True)
            except:
                pass
        
        # Garante que o arquivo existe e está completo
        if not os.path.exists(excel_path) or os.path.getsize(excel_path) == 0:
            raise Exception("Erro ao salvar arquivo Excel")
        
        # Espera um momento para garantir que o arquivo foi salvo completamente
        time.sleep(1)
        
        # Inicia a conversão para PDF em background
        pdf_path = os.path.join(app.config['DOWNLOAD_FOLDER'], pdf_filename)
        app.config['CONVERSION_QUEUE'].put((excel_path, pdf_path))
        
        return jsonify({
            'message': 'Arquivo gerado com sucesso',
            'excel_url': f'/download/{excel_filename}',
            'conversion_id': pdf_filename,
            'status_url': f'/conversion-status/{pdf_filename}'
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True)